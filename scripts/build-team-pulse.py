#!/usr/bin/env python3
"""
EDEN & CO. CEO Flight Deck — Team Pulse Cache Builder
Reads EDEN_CO_Weekly_Status.xlsx and writes src/data/team-pulse-cache.js

Usage (from build/ directory):
    python3 scripts/build-team-pulse.py

Run this each Monday morning before opening the dashboard, or whenever
the team has updated the spreadsheet.
"""

import json, os, sys, tempfile
import urllib.request
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print('ERROR: openpyxl not installed. Run: pip3 install openpyxl')
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BUILD_DIR  = os.path.dirname(SCRIPT_DIR)
LOCAL_PATH = os.path.join(BUILD_DIR, 'Final Data Files', 'Team Pulse', 'EDEN_CO_Weekly_Status.xlsx')
OUT_PATH   = os.path.join(BUILD_DIR, 'src', 'data', 'team-pulse-cache.js')

# Allow OneDrive share link override (for GitHub Actions)
ONEDRIVE_URL = os.environ.get('ONEDRIVE_XLSX_URL', '')
if ONEDRIVE_URL:
    download_url = ONEDRIVE_URL.replace('redir?', 'download?').replace('embed?', 'download?')
    if 'download' not in download_url:
        import base64
        encoded = base64.b64encode(ONEDRIVE_URL.encode()).decode().rstrip('=').replace('+','-').replace('/','_')
        download_url = f'https://api.onedrive.com/v1.0/shares/u!{encoded}/root/content'
    print(f'[INFO] Downloading XLSX from OneDrive...')
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    urllib.request.urlretrieve(download_url, tmp.name)
    XLSX_PATH = tmp.name
    print(f'[INFO] Downloaded to {tmp.name}')
else:
    XLSX_PATH = LOCAL_PATH

PEOPLE = ['Jon', 'Rosie', 'Edith', 'Phoebe']

def normalise(c):
    """Convert a cell value to a clean string, handling datetime objects."""
    if c is None:
        return ''
    from datetime import datetime as dt
    if isinstance(c, dt):
        return c.strftime('%Y-%m-%d')
    return str(c).strip()

def rows(ws, skip=3):
    """Yield dicts for each data row, skipping header rows.
    The header row is the skip-th row (1-indexed). Rows before it are
    title/instruction rows. Detects the header by looking for the first
    row that contains a known sentinel value if auto-detection is needed.
    """
    headers = None
    count = 0
    for row in ws.iter_rows(values_only=True):
        count += 1
        if count == skip:
            headers = [normalise(c) for c in row]
            continue
        if count <= skip:
            continue
        if headers and any(c is not None for c in row):
            yield dict(zip(headers, [normalise(c) for c in row]))


def build_week_key():
    from datetime import date
    n = date.today()
    return n.strftime('%G-W%V')


def main():
    if not os.path.exists(XLSX_PATH):
        print(f'ERROR: Sheet not found at {XLSX_PATH}')
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f'Opened: {XLSX_PATH}')
    print(f'Sheets: {wb.sheetnames}')

    # ── Core Goals ──────────────────────────────────────────────────
    core_goals = {}  # { Person: { goal_title: { status, subgoals: [{t, m}] } } }
    if 'Core Goals' in wb.sheetnames:
        ws = wb['Core Goals']
        for r in rows(ws, skip=2):
            person = r.get('Person', '').strip()
            title  = r.get('Goal Title', '').strip()
            sub    = r.get('Sub-goal / Metric', '').strip()
            target = r.get('Target', '').strip()
            status = r.get('Status', '').strip().lower()
            if not person or not title:
                continue
            if person not in core_goals:
                core_goals[person] = {}
            if title not in core_goals[person]:
                core_goals[person][title] = {'status': 'on' if status == 'active' else status, 'goals': []}
            if sub:
                core_goals[person][title]['goals'].append({'t': sub, 'm': target})
        print(f'Core Goals: {sum(len(v) for v in core_goals.values())} goals across {len(core_goals)} people')

    # ── Weekly actions (Jon, Rosie, Edith, Phoebe tabs) ─────────────
    weekly = {}  # { Person: [ {week, category, task, status, parent, note, other} ] }
    for person in PEOPLE:
        if person not in wb.sheetnames:
            print(f'  WARNING: No sheet for {person}')
            continue
        ws = wb[person]
        weekly[person] = []
        for r in rows(ws, skip=3):
            week  = r.get('Week', '').strip()
            task  = r.get('Task', '').strip()
            other = r.get('Other', '').strip()
            # Row with no task but an Other value = noise item
            if not task and other:
                weekly[person].append({
                    'week':     week,
                    'category': 'Other',
                    'task':     other,
                    'status':   '',
                    'parent':   '',
                    'note':     '',
                    'other':    '',
                })
                continue
            if not task:
                continue
            weekly[person].append({
                'week':     week,
                'category': r.get('Category', 'This week').strip() or 'This week',
                'task':     task,
                'status':   r.get('Status', '').strip().lower(),
                'parent':   r.get('Parent Goal', '').strip(),
                'note':     r.get('Note', '').strip(),
                'other':    other,
            })
        print(f'  {person}: {len(weekly[person])} action rows')

    # ── Concerns ────────────────────────────────────────────────────
    concerns = []
    if 'Concerns' in wb.sheetnames:
        ws = wb['Concerns']
        for r in rows(ws, skip=2):
            title  = r.get('Title', '').strip()
            status = r.get('Status', '').strip().lower()
            if not title or status == 'resolved':
                continue
            concerns.append({
                'title':  title,
                'owner':  r.get('Owner', '').strip(),
                'level':  r.get('Level', 'med').strip().lower(),
                'action': r.get('Action', '').strip(),
                'due':    r.get('Due Date', '').strip(),
            })
        print(f'Concerns: {len(concerns)} open items')

    # ── Assemble output ──────────────────────────────────────────────
    current_week = build_week_key()
    print(f'Current week key: {current_week}')

    # YTD done counts
    ytd = {p: 0 for p in PEOPLE}
    for person, actions in weekly.items():
        year = str(datetime.now().year)
        for a in actions:
            if a['week'].startswith(year) and a['status'] == 'done':
                ytd[person] += 1

    # Streak (distinct weeks with at least one done row)
    streak = {}
    for person, actions in weekly.items():
        streak[person] = len(set(a['week'] for a in actions if a['status'] == 'done'))

    output = {
        'generated':    datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        'current_week': current_week,
        'core_goals':   core_goals,
        'weekly':       weekly,
        'ytd':          ytd,
        'streak':       streak,
        'concerns':     concerns,
    }

    ts = output['generated']
    js = f"""// EDEN & CO. CEO Flight Deck — Team Pulse Cache
// Generated: {ts}
// Source: EDEN_CO_Weekly_Status.xlsx
// Do not edit manually. Run: python3 scripts/build-team-pulse.py

window.EDEN = window.EDEN || {{}};
window.EDEN.teamPulse = {json.dumps(output, indent=2, ensure_ascii=False)};
"""

    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js)

    print(f'\nWrote: {OUT_PATH}')
    print(f'YTD done: {ytd}')
    print(f'Streak weeks: {streak}')
    print('Done.')


if __name__ == '__main__':
    main()
